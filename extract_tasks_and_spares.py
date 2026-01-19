#!/usr/bin/env python
import argparse
import re
from copy import copy
from pathlib import Path
from typing import List, Dict, Tuple

from openpyxl import Workbook
import PyPDF2


# ----------------------------------------------------------------------
# PDF → lines (no OCR, pure text)
# ----------------------------------------------------------------------
def pdf_to_text_lines(pdf_path: Path) -> List[str]:
    """
    Read the PDF with PyPDF2 and return a flat list of non-empty text lines.
    Page footers like 'Database:' or 'Printed by' are filtered out.
    """
    reader = PyPDF2.PdfReader(str(pdf_path))
    all_lines: List[str] = []

    for page in reader.pages:
        text = page.extract_text() or ""
        if not text.strip():
            continue

        for raw_ln in text.splitlines():
            ln = raw_ln.rstrip()
            if not ln.strip():
                continue

            low = ln.lower().strip()
            if low.startswith("database:") or low.startswith("printed by"):
                continue

            all_lines.append(ln)

    return all_lines


# ----------------------------------------------------------------------
# Helpers & heuristics
# ----------------------------------------------------------------------
def is_header_line(line: str) -> bool:
    """
    Detect the main Task header line:
    'Task Code Trade Task Action Task Description Doc Ref Interval'
    """
    lower = " ".join(line.lower().split())
    return "task code" in lower and "task action" in lower


def is_spares_header_line(line: str) -> bool:
    """
    Detect the Spare parts header:
    'Part No Part Description Task Code Task Action Component Tree Path Qty Required Unit Of Measure'
    """
    lower = " ".join(line.lower().split())
    return "part no" in lower and "task code" in lower and "qty required" in lower


def is_metadata_line(line: str) -> bool:
    """
    Filter out page/meta lines.
    """
    low = line.lower().strip()
    return low.startswith("database:") or low.startswith("printed by") or low.startswith("page ")


def parse_asset_line(line: str) -> Tuple[str, str]:
    """
    'Asset: 9000171371 TP A3/F-040V ...' → (asset_code, asset_type_string)
    """
    parts = line.split(":", 1)[1].strip().split()
    if not parts:
        return "", ""
    asset_code = parts[0]
    asset_type = " ".join(parts[1:]) if len(parts) > 1 else ""
    return asset_code, asset_type


def looks_like_component_line(line: str) -> bool:
    """
    Grey rows with component / location info:
    e.g. '1 Pre-Maintenance Checks: (9000171371) ... \ [648575-0400] ...'
    """
    stripped = line.strip()
    if not stripped:
        return False
    if "\\" in line or "[" in line or stripped.startswith("("):
        return True
    if re.match(r"^\(\d{6,}-\d{4,}:", stripped):
        return True
    return False


def parse_grey_row(line: str):
    """
    From a grey row line, extract:
    - Location1 (before '\')
    - Location2 (after '\')
    - setTypeCode (digits inside parentheses, last occurrence)
    - ComponentPath (full raw line)
    """
    component_path = line.strip()
    if "\\" in line:
        a, b = line.split("\\", 1)
    else:
        a, b = line, ""
    location1 = a.strip()
    location2 = b.strip()
    codes = re.findall(r"\((\d+)\)", line)
    set_type_code = codes[-1] if codes else ""
    return location1, location2, set_type_code, component_path


def strip_status_prefix(line: str) -> str:
    """
    Remove bullets / markers before the first digit or '*'.
    """
    m = re.search(r"[\*\d]", line)
    return line[m.start():].strip() if m else line.strip()


def looks_like_task_row(line: str) -> bool:
    """
    Detect rows that begin a Task:
    '*9465150 ENGR Check Check ...'
    """
    if is_metadata_line(line):
        return False
    if looks_like_component_line(line) and " " not in strip_status_prefix(line).split()[0]:
        return False

    tokens = strip_status_prefix(line).split()
    if len(tokens) < 3:
        return False

    code_token = tokens[0]
    if "/" in code_token:
        return False

    if not re.match(r"\*?\d{6,8}$", code_token):
        return False

    trade_token = tokens[1]
    return trade_token.isalpha() and trade_token.isupper()


def normalize_task_code(code: str) -> str:
    """
    Remove leading '*' for consistent lookup.
    """
    return re.sub(r"^\*", "", code)


def gather_task_block(lines: List[str], idx: int) -> Tuple[str, int]:
    """
    Given index of a line that starts a task, gather continuation lines
    until we hit a new task / component / asset / header / metadata.
    Returns (combined_text, next_index).
    """
    buf = [lines[idx]]
    i = idx + 1
    n = len(lines)

    while i < n:
        ln = lines[i]
        if not ln.strip():
            i += 1
            continue
        if is_metadata_line(ln) or is_header_line(ln):
            break
        if looks_like_task_row(ln) or looks_like_component_line(ln) or ln.lower().startswith("asset:"):
            break
        buf.append(ln)
        i += 1

    return " ".join(buf), i


def parse_task_row(full_line: str):
    """
    Parse a full task row (possibly assembled from multiple lines).

    Handles patterns like:
      '*9465150 ENGR Check Check ... No reference 1000 Hours'
      '*9465160 TECH Check Check Warning labels MM 1000 Hours'
    """
    normalized = strip_status_prefix(full_line)
    tokens = normalized.split()
    if len(tokens) < 3:
        return "", "", "", "", "", ""

    task_code, trade, task_action = tokens[0], tokens[1], tokens[2]
    rest = " ".join(tokens[3:]).strip()

    # Pattern 1: description + 'No reference 1000 Hours'
    m = re.search(
        r"^(?P<desc>.*)\bNo reference\b\s+(?P<num>\d+)\s+(?P<unit>Hours?|Weeks?|Months?|Days?)\b$",
        rest,
    )
    if m:
        desc = m.group("desc").strip()
        return task_code, trade, task_action, desc, "No reference", f"{m.group('num')} {m.group('unit')}"

    # Pattern 2: description + DOCREF + '1000 Hours'
    m = re.search(
        r"^(?P<desc>.*)\s(?P<doc>[A-Z0-9./-]{1,10})\s+(?P<num>\d+)\s+(?P<unit>Hours?|Weeks?|Months?|Days?)\b$",
        rest,
    )
    if m:
        desc = m.group("desc").strip()
        doc_ref = m.group("doc").strip()
        return task_code, trade, task_action, desc, doc_ref, f"{m.group('num')} {m.group('unit')}"

    # Pattern 3: description + DOCREF + 'No Interval'
    m = re.search(
        r"^(?P<desc>.*)\s(?P<doc>[A-Z0-9./-]{1,10})\s+No Interval$",
        rest,
        flags=re.IGNORECASE,
    )
    if m:
        desc = m.group("desc").strip()
        doc_ref = m.group("doc").strip()
        return task_code, trade, task_action, desc, doc_ref, "No Interval"

    # Fallback: best-effort split
    remaining_tokens = rest.split()

    def _split_interval(tokens):
        if not tokens:
            return [], ""
        lowers = [t.lower() for t in tokens]
        if len(tokens) >= 2 and lowers[-2:] == ["no", "interval"]:
            return tokens[:-2], "No Interval"
        if (
            lowers[-1] in {"hours", "hour", "week", "weeks", "month", "months", "day", "days"}
            and len(tokens) >= 2
            and re.fullmatch(r"\d+", tokens[-2])
        ):
            return tokens[:-2], f"{tokens[-2]} {tokens[-1]}"
        return tokens, ""

    def _split_doc_ref_and_desc(tokens):
        if not tokens:
            return "", ""
        lowers = [t.lower() for t in tokens]
        # 'No reference' anywhere near the end
        for i in range(len(tokens) - 1):
            if lowers[i] == "no" and lowers[i + 1] == "reference":
                return " ".join(tokens[:i]).strip(), "No reference"
        # Otherwise, doc ref is trailing upper-case / numeric chunk
        lookback = tokens[-5:] if len(tokens) > 5 else tokens
        start = len(tokens) - 1
        for offset, tok in enumerate(lookback):
            if re.search(r"\d", tok) or re.search(r"[:;/.\-]", tok) or (tok.isupper() and len(tok) <= 4):
                start = len(tokens) - len(lookback) + offset
                break
        return " ".join(tokens[:start]).strip(), " ".join(tokens[start:]).strip()

    body, interval = _split_interval(remaining_tokens)
    desc, doc_ref = _split_doc_ref_and_desc(body)
    return task_code, trade, task_action, desc, doc_ref, interval


def looks_like_part_line(line: str) -> bool:
    """
    Detect a spare part row by the '1234567-0000' style part number.
    """
    return bool(re.search(r"\b\d{6,}-\d{3,}\b", line))


def gather_part_block(lines: List[str], idx: int) -> Tuple[str, int]:
    """
    Gather one spare part record that may span multiple lines.
    """
    buf = [lines[idx]]
    i = idx + 1
    n = len(lines)

    while i < n:
        nxt = lines[i]
        if not nxt.strip():
            i += 1
            continue
        if is_spares_header_line(nxt) or is_metadata_line(nxt) or nxt.lower().startswith("asset:"):
            break
        if looks_like_part_line(nxt) and i != idx:
            break
        buf.append(nxt)
        i += 1

    return " ".join(buf), i


def parse_part_block(lines: List[str], idx: int):
    """
    Parse one spare part block starting at lines[idx].
    Returns (dict_or_None, next_index).
    """
    combined, next_idx = gather_part_block(lines, idx)
    tokens = combined.split()
    if not tokens:
        return None, next_idx

    # Part number
    part_idx = None
    for j, tok in enumerate(tokens):
        if re.fullmatch(r"\d{6,}-\d{3,}", tok):
            part_idx = j
            break
    if part_idx is None:
        return None, next_idx

    part_no = tokens[part_idx]

    # Task code
    task_idx = None
    for j in range(part_idx + 1, len(tokens)):
        if re.fullmatch(r"\*?\d{6,8}", tokens[j]):
            task_idx = j
            break

    task_code = normalize_task_code(tokens[task_idx]) if task_idx is not None else ""
    task_action = tokens[task_idx + 1] if task_idx is not None and task_idx + 1 < len(tokens) else ""

    # Part description (between part number and task code)
    desc_tokens = tokens[part_idx + 1 : task_idx if task_idx else len(tokens)]

    # Component path + qty + UOM (after task action)
    comp_tokens = tokens[(task_idx + 2) if task_idx is not None else len(tokens) :]

    uom = ""
    qty = ""

    if comp_tokens:
        last = comp_tokens[-1]
        if re.fullmatch(r"[A-Za-z]{1,4}", last):
            uom = last
            comp_tokens = comp_tokens[:-1]

    if comp_tokens:
        last = comp_tokens[-1]
        if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", last):
            qty = last
            comp_tokens = comp_tokens[:-1]

    component_path = " ".join(comp_tokens).strip()
    part_description = " ".join(desc_tokens).strip()

    return (
        {
            "TaskCode": task_code,
            "TaskAction": task_action,
            "PartNo": part_no,
            "PartDescription": part_description,
            "QtyRequired": qty,
            "UOM": uom,
            "ComponentPath": component_path,
        },
        next_idx,
    )


# ----------------------------------------------------------------------
# Extraction logic
# ----------------------------------------------------------------------
def extract_tasks(pdf_path: Path):
    """
    Extract all tasks from the whole PDF.
    Returns:
        task_rows: list of row dicts for Tasks sheet
        task_lookup: dict TaskCode -> row dict
        asset_type, asset_code, lines (for spares)
    """
    lines = pdf_to_text_lines(pdf_path)

    rows = []
    rows_by_code: Dict[str, Dict] = {}
    current_loc1 = ""
    current_loc2 = ""
    current_setcode = ""
    current_comppath = ""
    asset_type = ""
    asset_code = ""

    i = 0
    n = len(lines)

    while i < n:
        ln = lines[i]
        low = ln.lower()

        if low.startswith("asset:"):
            asset_code, asset_type = parse_asset_line(ln)
            i += 1
            continue

        if is_header_line(ln):
            i += 1
            continue

        if is_metadata_line(ln):
            i += 1
            continue

        if looks_like_component_line(ln):
            current_loc1, current_loc2, current_setcode, current_comppath = parse_grey_row(ln)
            i += 1
            continue

        if looks_like_task_row(ln):
            block, next_i = gather_task_block(lines, i)
            task_code, trade, action, desc, doc_ref, interval = parse_task_row(block)
            norm = normalize_task_code(task_code)

            row = {
                "TaskCode": norm,
                "Trade": trade,
                "TaskAction": action,
                "TaskDescription": desc,
                "DocRef": doc_ref,
                "Interval": interval,
                "Location1": current_loc1,
                "Location2": current_loc2,
                "setTypeCode": current_setcode,
                "ComponentPath": current_comppath,
                # Extra columns (blank or defaults)
                "TypeOfWork": "",
                "MotionType": "",
                "Duration": "",
                "DurationCalc": "",
                "DurationUOM": "",
                "MTBMPredicted": "",
                "CostCode": "",
                "IncludeInME": "",
                "TaskDependency": "",
                "FollowUpTasks": "",
                "LocationDependency": "",
                "Active": "Y",
                "Section": "",
                "AssetType": asset_type,
                "AssetTypeCode": current_setcode or asset_code,
            }

            if norm in rows_by_code:
                existing = rows_by_code[norm]
                # merge: prefer longer description, fill missing fields
                if len(desc) > len(existing.get("TaskDescription", "")):
                    existing["TaskDescription"] = desc
                if not existing.get("DocRef") and doc_ref:
                    existing["DocRef"] = doc_ref
                if not existing.get("Interval") and interval:
                    existing["Interval"] = interval
                if not existing.get("Location1") and current_loc1:
                    existing["Location1"] = current_loc1
                if not existing.get("Location2") and current_loc2:
                    existing["Location2"] = current_loc2
            else:
                rows.append(row)
                rows_by_code[norm] = row

            i = next_i
            continue

        i += 1

    return rows, rows_by_code, asset_type, asset_code, lines


def extract_spares_from_lines(lines: List[str], task_lookup: Dict[str, Dict]):
    """
    Extract all spare parts from the lines, associating them with the task
    context (TaskCode, locations, setTypeCode) whenever possible.
    """
    spare_rows = []
    current_task_code = ""
    current_loc1 = ""
    current_loc2 = ""
    current_setcode = ""
    current_comppath = ""
    seen_keys = set()

    i = 0
    n = len(lines)

    while i < n:
        ln = lines[i]

        if is_metadata_line(ln):
            i += 1
            continue

        low = ln.lower()
        if low.startswith("asset:"):
            i += 1
            continue

        if is_spares_header_line(ln):
            i += 1
            continue

        if looks_like_component_line(ln):
            current_loc1, current_loc2, current_setcode, current_comppath = parse_grey_row(ln)
            i += 1
            continue

        if looks_like_task_row(ln):
            code = normalize_task_code(strip_status_prefix(ln).split()[0])
            current_task_code = code
            i += 1
            continue

        if looks_like_part_line(ln):
            parsed, next_i = parse_part_block(lines, i)
            i = next_i
            if not parsed:
                continue

            task_code = parsed.get("TaskCode") or current_task_code
            if not task_code:
                continue

            task_ctx = task_lookup.get(task_code, {})
            key = (task_code, parsed["PartNo"], parsed["PartDescription"])
            if key in seen_keys:
                continue
            seen_keys.add(key)

            loc1 = current_loc1 or task_ctx.get("Location1", "")
            loc2 = current_loc2 or task_ctx.get("Location2", "")
            setcode = current_setcode or task_ctx.get("setTypeCode", "")
            if not setcode:
                matches = re.findall(r"\((\d+)\)", parsed.get("ComponentPath", ""))
                if matches:
                    setcode = matches[-1]

            spare_rows.append(
                {
                    "TaskCode": task_code,
                    "PartNo": parsed["PartNo"],
                    "PartDescription": parsed["PartDescription"],
                    "MU_TL": "",
                    "QtyRequired": parsed["QtyRequired"],
                    "UOM": parsed["UOM"],
                    "ItemDependency": "",
                    "Location1": loc1,
                    "Location2": loc2,
                    "AssetType": "",
                    "AssetTypeCode": setcode,
                }
            )
            continue

        i += 1

    return spare_rows


# ----------------------------------------------------------------------
# Workbook builder
# ----------------------------------------------------------------------
def build_workbook(task_rows: List[Dict], spare_rows: List[Dict]) -> Workbook:
    wb = Workbook()

    # --- Tasks sheet ---
    ws_tasks = wb.active
    ws_tasks.title = "Tasks"
    ws_tasks.delete_rows(1, ws_tasks.max_row)

    task_headers = [
        "Sort",
        "TaskCode",
        "TaskAction",
        "TaskDescription",
        "TypeOfWork",
        "MotionType",
        "Duration",
        "DurationCalc",
        "DurationUOM",
        "Interval",
        "MTBMPredicted",
        "CostCode",
        "IncludeInME",
        "TaskDependency",
        "FollowUpTasks",
        "LocationDependency",
        "Active",
        "Trade",
        "Section",
        "DocRef",
        "Location1",
        "Location2",
        "ComponentPath",
        "AssetType",
        "AssetTypeCode",
    ]

    ws_tasks.append(task_headers)
    for cell in ws_tasks[1]:
        bold_font = copy(cell.font)
        bold_font.bold = True
        cell.font = bold_font
    ws_tasks.row_dimensions[1].height = 22

    for idx, row in enumerate(task_rows, start=1):
        row["Sort"] = idx
        ws_tasks.append([row.get(h, "") for h in task_headers])

    # --- SpareParts sheet ---
    ws_spares = wb.create_sheet("SpareParts")
    spare_headers = [
        "TaskCode",
        "PartNo",
        "PartDescription",
        "MU_TL",
        "QtyRequired",
        "UOM",
        "ItemDependency",
        "Location1",
        "Location2",
        "AssetType",
        "AssetTypeCode",
    ]

    ws_spares.append(spare_headers)
    for cell in ws_spares[1]:
        bold_font = copy(cell.font)
        bold_font.bold = True
        cell.font = bold_font
    ws_spares.row_dimensions[1].height = 22

    for row in spare_rows:
        ws_spares.append([row.get(h, "") for h in spare_headers])

    return wb


# ----------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Extract Tasks and Spare Parts from a text-based maintenance PDF into Excel."
    )
    parser.add_argument(
        "--pdf",
        required=True,
        help="Path to input PDF file.",
    )
    parser.add_argument(
        "--out",
        help="Output XLSX path. Defaults to <pdf_stem>_tasks_spares.xlsx",
    )

    args = parser.parse_args()
    pdf_path = Path(args.pdf)

    if args.out:
        output_xlsx = Path(args.out)
    else:
        output_xlsx = pdf_path.with_name(f"{pdf_path.stem}_tasks_spares.xlsx")

    print(f"Reading PDF: {pdf_path}")
    task_rows, task_lookup, asset_type, asset_code, lines = extract_tasks(pdf_path)
    print(f"Extracted {len(task_rows)} task rows.")
    spare_rows = extract_spares_from_lines(lines, task_lookup)
    print(f"Extracted {len(spare_rows)} spare part rows.")

    wb = build_workbook(task_rows, spare_rows)
    wb.save(output_xlsx)
    print(f"Saved Excel file: {output_xlsx}")


if __name__ == "__main__":
    main()
